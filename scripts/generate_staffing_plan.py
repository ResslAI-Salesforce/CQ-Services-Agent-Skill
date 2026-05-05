"""
CQ Services Workflow — Staffing Plan Generator
================================================
Populates the bundled Flexshare template (two sheets: GxP + Non-GxP).
Only the relevant sheet is populated based on is_gxp flag.

Usage:
    python generate_staffing_plan.py '<json_input>' output.xlsx

JSON input schema:
{
  "customer_name": "Acme Corp",
  "processes": ["Document", "Training", "Change"],
  "is_gxp": false,
  "roles": {
    "cc_hours": 380,
    "sc_hours": 190,
    "sa_hours": 76,
    "pm_hours": 76,
    "ta_hours": 0,
    "validation_hours": 0
  },
  "hypercare_hours": 40,
  "rollout_hours": 0
}
"""

import sys, json, math, shutil
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

TEMPLATE_PATH = Path(__file__).parent.parent / "assets" / "staffing-plan-template.xlsx"

RATES = {
    "pm": 150, "sa": 125, "ta": 150,
    "sc": 125, "cc": 60, "validation": 125,
    "hypercare": 125, "rollout": 125,
}

ROLE_NAMES = {
    "pm": "Project Manager", "sa": "Solution Architect",
    "ta": "Technical Architect", "sc": "Solution Consultant",
    "cc": "Configuration Consultants", "validation": "Validation Specialist",
    "hypercare": "Hypercare", "rollout": "Rollout Support",
}

ROLE_ORDER = ["pm", "sa", "ta", "sc", "cc", "validation", "hypercare", "rollout"]


def compute_phases(cc_hours, hypercare_hours):
    """Phase week counts mirroring the Flexshare template structure."""
    n   = math.ceil(cc_hours / 40)           # config core weeks at 1 FTE
    ci1 = max(2, math.ceil(n * 0.40))        # Config Iteration 1
    ci2 = max(2, n - ci1)                    # Config Iteration 2
    uat = max(2, math.ceil(n * 0.35))        # UAT / TTT
    hyp = max(1, math.ceil(hypercare_hours / 10)) if hypercare_hours > 0 else 0
    return {"prep": 1, "workshop": 1, "ci1": ci1, "ci2": ci2,
            "final_config": 1, "uat": uat, "golive": 1, "hypercare": hyp}


def phase_offsets(phases):
    """0-indexed (start, length) per phase, plus total."""
    order = ["prep", "workshop", "ci1", "ci2", "final_config", "uat", "golive", "hypercare"]
    offsets, pos = {}, 0
    for k in order:
        offsets[k] = (pos, phases[k])
        pos += phases[k]
    offsets["total"] = pos
    return offsets


def distribute(total, n):
    """Spread total evenly across n slots, integer output."""
    if not n or not total:
        return [0] * (n or 0)
    base, rem = divmod(int(round(total)), n)
    return [base + (1 if i < rem else 0) for i in range(n)]


def fill_phase(empty_schedule, offsets, phase_keys, hours):
    """Place hours across named phases."""
    sched = list(empty_schedule)
    if not hours:
        return sched
    weeks = []
    for pk in phase_keys:
        start, length = offsets[pk]
        weeks.extend(range(start, start + length))
    dist = distribute(hours, len(weeks))
    for i, wi in enumerate(weeks):
        sched[wi] = dist[i]
    return sched


def build_schedules(roles, offsets, is_gxp, hypercare_hours, rollout_hours):
    n = offsets["total"]
    empty = lambda: [0] * n

    # PM: steady across all active phases (not hypercare)
    pm = empty()
    active = list(range(offsets["hypercare"][0]))
    if roles["pm_hours"] and active:
        for i, wi in enumerate(active):
            pm[wi] = distribute(roles["pm_hours"], len(active))[i]

    # SA: front-loaded — 50% prep+workshop, 40% ci1, 10% ci2
    sa = empty()
    if roles["sa_hours"]:
        t = roles["sa_hours"]
        ps, pl = offsets["prep"];      early_w  = list(range(ps, ps + pl))
        ws2, wl = offsets["workshop"]; early_w += list(range(ws2, ws2 + wl))
        cs, cl = offsets["ci1"];       ci1_w    = list(range(cs, cs + cl))
        c2s, c2l = offsets["ci2"];     ci2_w    = list(range(c2s, c2s + c2l))
        e_hrs, ci1_hrs = round(t * 0.50), round(t * 0.40)
        ci2_hrs = t - e_hrs - ci1_hrs
        for wi, h in zip(early_w, distribute(e_hrs,   len(early_w))): sa[wi] = h
        for wi, h in zip(ci1_w,   distribute(ci1_hrs, len(ci1_w))):   sa[wi] = h
        for wi, h in zip(ci2_w,   distribute(ci2_hrs, len(ci2_w))):   sa[wi] = h

    # TA: spread across ci1 + ci2 + uat (integration work)
    ta = fill_phase(empty(), offsets, ["ci1", "ci2", "uat"], roles.get("ta_hours", 0))

    # SC: active from workshop through uat
    sc = fill_phase(empty(), offsets, ["workshop", "ci1", "ci2", "final_config", "uat"],
                    roles["sc_hours"])

    # CC: ci1 + ci2 + final_config (max 40/week enforced by compute_phases)
    cc = fill_phase(empty(), offsets, ["ci1", "ci2", "final_config"], roles["cc_hours"])

    # Validation (GxP only): 25% in ci2, 75% in uat
    val = empty()
    if is_gxp and roles.get("validation_hours", 0):
        vh = roles["validation_hours"]
        ci2_hrs, uat_hrs = round(vh * 0.25), round(vh * 0.75)
        uat_hrs += vh - ci2_hrs - uat_hrs   # absorb rounding
        for phase_key, hrs in [("ci2", ci2_hrs), ("uat", uat_hrs)]:
            start, length = offsets[phase_key]
            for i, wi in enumerate(range(start, start + length)):
                val[wi] = distribute(hrs, length)[i]

    hyp = fill_phase(empty(), offsets, ["hypercare"], hypercare_hours)
    rol = fill_phase(empty(), offsets, ["hypercare"], rollout_hours)

    return {"pm": pm, "sa": sa, "ta": ta, "sc": sc,
            "cc": cc, "validation": val, "hypercare": hyp, "rollout": rol}


def safe_unmerge_row(ws, row, col_start, col_end):
    for merge in list(ws.merged_cells.ranges):
        if merge.min_row <= row <= merge.max_row:
            if not (merge.max_col < col_start or merge.min_col > col_end):
                try:
                    ws.unmerge_cells(start_row=merge.min_row, start_column=merge.min_col,
                                     end_row=merge.max_row, end_column=merge.max_col)
                except Exception:
                    pass


def write_sheet(ws, is_gxp, schedules, phases, offsets,
                customer_name, processes, n_weeks):
    FIRST_COL      = 6
    TEMPLATE_WEEKS = 18

    # Row layout differs between GxP and Non-GxP
    if is_gxp:
        val_band_row = 4; phase_row = 5; header_row = 6
        first_data = 7;   total_row  = 15
    else:
        val_band_row = None; phase_row = 4; header_row = 5
        first_data = 6;      total_row  = 14

    last_col = FIRST_COL + n_weeks - 1

    # Adjust column count
    if n_weeks > TEMPLATE_WEEKS:
        ws.insert_cols(FIRST_COL + TEMPLATE_WEEKS, n_weeks - TEMPLATE_WEEKS)
    elif n_weeks < TEMPLATE_WEEKS:
        ws.delete_cols(FIRST_COL + n_weeks, TEMPLATE_WEEKS - n_weeks)

    # Process header
    proc_str = ", ".join(processes[:4]) or "To be confirmed"
    ws["B1"] = f"{customer_name} — {proc_str}"
    ws["B2"] = f"{len(processes)} process(es) in scope"

    # Week labels (W1, W2, ...)
    safe_unmerge_row(ws, header_row, FIRST_COL, last_col)
    for i in range(n_weeks):
        c = ws.cell(row=header_row, column=FIRST_COL + i)
        c.value = f"W{i+1}"
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=True, size=11)

    # Phase band labels
    safe_unmerge_row(ws, phase_row, FIRST_COL, last_col)

    phase_labels = [
        ("prep",         "Prep",                             False, False),
        ("workshop",     "Workshop",                         False, False),
        ("ci1",          "Config Iteration 1",               False, False),
        ("ci2",          "Config Iteration 2",               False, False),
        ("final_config", "Final Config",                     False, False),
        ("uat",          "Train the Trainer Prep & TTT" if is_gxp
                         else "UAT Support / Train the Trainer Prep & TTT",
                                                             False, False),
        ("golive",       "Go-live",                          True,  True),
        ("hypercare",    "Hypercare",                        False, False),
    ]

    for pk, label, red, bold in phase_labels:
        start, length = offsets[pk]
        col = FIRST_COL + start
        c = ws.cell(row=phase_row, column=col)
        c.value = label
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=bold, color="FF0000" if red else "000000", size=11)
        if length > 1:
            try:
                ws.merge_cells(start_row=phase_row, start_column=col,
                               end_row=phase_row, end_column=col + length - 1)
            except Exception:
                pass

    # Config lock — clear entire row 3 first, then place once at Final Config column
    safe_unmerge_row(ws, 3, FIRST_COL, last_col)
    for col in range(FIRST_COL, last_col + 1):
        try:
            ws.cell(row=3, column=col).value = None
        except Exception:
            pass
    lock_col = FIRST_COL + offsets["final_config"][0]
    c = ws.cell(row=3, column=lock_col)
    c.value = "Config lock"
    c.font = Font(bold=True, color="FF0000", size=11)
    c.alignment = Alignment(horizontal="center")

    # Validation band (GxP only)
    if val_band_row:
        safe_unmerge_row(ws, val_band_row, FIRST_COL, last_col)
        v_start = FIRST_COL + offsets["ci2"][0]
        v_end   = FIRST_COL + offsets["uat"][0] + offsets["uat"][1] - 1
        c = ws.cell(row=val_band_row, column=v_start)
        c.value = "Validation"
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=True, size=11)
        if v_end > v_start:
            try:
                ws.merge_cells(start_row=val_band_row, start_column=v_start,
                               end_row=val_band_row, end_column=v_end)
            except Exception:
                pass

    # Role data rows
    for i, role_key in enumerate(ROLE_ORDER):
        row  = first_data + i
        sched = schedules.get(role_key, [0] * n_weeks)
        last_letter = get_column_letter(last_col)
        ws.cell(row=row, column=1).value = ROLE_NAMES[role_key]
        ws.cell(row=row, column=2).value = f"=SUM(F{row}:{last_letter}{row})"
        ws.cell(row=row, column=3).value = RATES[role_key]
        ws.cell(row=row, column=4).value = f"=B{row}*C{row}"
        ws.cell(row=row, column=5).value = ROLE_NAMES[role_key]
        for j, hrs in enumerate(sched[:n_weeks]):
            ws.cell(row=row, column=FIRST_COL + j).value = int(round(hrs))

    # Total row
    ws.cell(row=total_row, column=1).value = "Total"
    ws.cell(row=total_row, column=4).value = \
        f"=SUM(D{first_data}:D{first_data + len(ROLE_ORDER) - 1})"


def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_staffing_plan.py '<json>' output.xlsx")
        sys.exit(1)

    params          = json.loads(sys.argv[1])
    output_path     = sys.argv[2]
    roles           = params["roles"]
    is_gxp          = params.get("is_gxp", False)
    hypercare_hours = params.get("hypercare_hours", 40)
    rollout_hours   = params.get("rollout_hours", 0)

    phases    = compute_phases(roles["cc_hours"], hypercare_hours)
    offsets   = phase_offsets(phases)
    n_weeks   = offsets["total"]
    schedules = build_schedules(roles, offsets, is_gxp, hypercare_hours, rollout_hours)

    shutil.copy(TEMPLATE_PATH, output_path)
    wb = openpyxl.load_workbook(output_path)
    sheet_name = "Per Phase (GxP)" if is_gxp else "Per Phase (Non-GxP)"
    write_sheet(wb[sheet_name], is_gxp, schedules, phases, offsets,
                params.get("customer_name", "Customer"),
                params.get("processes", []), n_weeks)
    wb.save(output_path)

    print(json.dumps({
        "output": output_path, "sheet": sheet_name,
        "total_weeks": n_weeks, "phases": phases,
        "role_totals": {r: sum(s) for r, s in schedules.items()},
    }, indent=2))


if __name__ == "__main__":
    main()
